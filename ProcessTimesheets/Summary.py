#import os
import datetime
import logging
from   Calendar  import Calendar
from   TimeSheet import Timesheet
import openpyxl
from   openpyxl.workbook import Workbook
from   openpyxl.styles   import Font,Side,Border,Alignment,Color,Style
from   openpyxl.styles.colors import RED
from   FAETeam import FAETeam
from   Metrics import Metrics


#-----------------------------------------------------------------------
# Class TSSummary
#-----------------------------------------------------------------------

class TSSummary:

  colName = \
    [
      'METRICS',
      '  Code Based',
      '    Key Accounts',
      '      Ericsson',
      '      Nokia',
      '      ALU',
      '      All Others',
      '    Regional Key Accounts',
      '      Qualcomm',
      '      AT&T',
      '      Sprint',
      '    Carriers',
      '      AT&T',
      '      Sprint',
      '      T-Mobile',
      '    Small Cell',
      '      Qualcomm',
      '      Intel',
      '      Parallel',
      '      SpiderCloud',
      '    Modular Instruments',
      '      Qor',
      '      Ter',
      '    Semiconductor',
      '      Air',
      '      Van',
      '    Others',
      '      OTH',
      '      COB',
      '      TTT',
      '    Overhead',
      '      X4x',
      '      X1x',
      '  Team Based',
      '    DMR',
      '    MI',
      '  Total Hours',
      '    All',
      '    Permenent',
      '    Contract',
      '    Labour',
      '    Travel',
      '    Standby',
      '  FAEs'
    ]


  def __init__(self,fname):
    self.filename = fname
    self.sslist  = []         # list of spreadsheets by week
    self.tsdict  = {}         # dict of timesheet data by week
    self.week    = None       # current week number
    self.year    = None       # year of query
    self.swb     = None       # summary workbook
    self.metrics = {}         # summary metrics
     
  #---------------------------------------------------------------------
  def Process(self,tsdata,year,week,region,team):
    self.year = year
    self.week = week

    self.Sort(tsdata,week)

    for i,week in enumerate(self.sslist):
      wsDate = Calendar.week[i+1]
      self.tsdict[wsDate] = []
      for j,ssdata in enumerate(week):
        ts = Timesheet(ssdata, wsDate)
        ts.ReadFile()
        self.tsdict[wsDate].append(ts)
 
    self.createWorkbook(region)
    self.writeRawDataSheet(region,team)
    self.saveWorkbook()

  #---------------------------------------------------------------------
  def createWorkbook(self,region):
    self.swb = Workbook()
    self.swb.create_sheet(region + ' Charts')
    self.swb.create_sheet(region + ' Tables')
    self.swb.create_sheet(region + ' Metrics')
    self.swb.create_sheet(region + ' Data')
    self.swb.remove_sheet(self.swb.get_sheet_by_name('Sheet'))

  #---------------------------------------------------------------------
  def saveWorkbook(self):
    self.swb.save(self.filename)

  #---------------------------------------------------------------------
  def setCell(self,cell,align,fmt,value):
    if (align == 'C'):
      align = Alignment(horizontal='center',vertical='center')
    elif (align == 'L'):
      align = Alignment(horizontal='left',vertical='center')
    elif (align == 'R'):
      align = Alignment(horizontal='right',vertical='center')
    else:
      align = Alignment(horizontal='right',vertical='center')


    side   = Side(style='thin')
    border = Border(left=side,right=side,top=side,bottom=side)
    #style  = Style(border=border,alignment=align,number_format=fmt)
    #cell.style = style

    if (fmt == 'F'):
      fmt = '0.00'
      cell.number_format = fmt
    cell.alignment     = align.copy()
    cell.border        = border.copy()
    cell.value = value

  #---------------------------------------------------------------------
  def flagCell(self,cell):
    side   = Side(style='medium',color=RED)
    border = Border(left=side,right=side,top=side,bottom=side)
    #style  = Style(border=border)
    #cell.style = style
    cell.border = border.copy()

  #---------------------------------------------------------------------
  def writeRawDataSheet(self,region,team):
    ws_name = region + ' Data'
    ws = self.swb.get_sheet_by_name(ws_name)

    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width =  3
    ws.column_dimensions['D'].width =  5
    ws.column_dimensions['E'].width =  3
    ws.column_dimensions['F'].width = 11
    ws.column_dimensions['G'].width = 11
    ws.column_dimensions['H'].width =  7
    ws.column_dimensions['I'].width =  5
    ws.column_dimensions['J'].width =  5
    ws.column_dimensions['K'].width =  5
    ws.column_dimensions['L'].width =  8
    ws.column_dimensions['M'].width = 10
    ws.column_dimensions['N'].width = 80

    wsRow = 2
    wsCol = 2

    #-------------------------------------------------------------------
    # Process each week
    #-------------------------------------------------------------------
    for weekIndex in range(1,self.week+1): 
      wsDate = Calendar.week[weekIndex]
      tslist = self.tsdict[wsDate]
      metrics = Metrics(wsDate)
      #-----------------------------------------------------------------
      # Process timesheet specific informaton
      #-----------------------------------------------------------------
      for ts in tslist:
        fae = ts.ssdata.fae

        #---------------------------------------------------------------
        # Process each entry in the timesheet
        #---------------------------------------------------------------
        for entry in ts.entries:
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 0),'L','G',fae.fullname.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 1),'C','G',fae.laborType.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 2),'C','G',fae.team.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 3),'C','G',fae.loc.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 4),'C','G',ts.ssdata.wsDate.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 5),'C','D',entry.date.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 6),'C','G',entry.code.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 7),'C','G',entry.location.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 8),'C','G',entry.activity.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+ 9),'C','G',entry.product.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+10),'R','F',entry.hours.GetWSVal())
          self.setCell(ws.cell(row=wsRow,column=wsCol+11),'C','G',entry.workType.GetWSVal())

          # Flag as red things to check
          if (entry.code.GetVal() == 'OTH'):
            self.flagCell(ws.cell(row=wsRow,column=wsCol+ 6))
            self.setCell(ws.cell(row=wsRow,column=wsCol+12),'L','G',entry.note.GetWSVal())
            self.flagCell(ws.cell(row=wsRow,column=wsCol+12))

          if (entry.code.GetVal() == 5):
            if (entry.location.GetVal() != None):
              self.flagCell(ws.cell(row=wsRow,column=wsCol+ 7))
            if (entry.activity.GetVal() != None):
              self.flagCell(ws.cell(row=wsRow,column=wsCol+ 8))
            if (entry.product.GetVal() != None):
              self.flagCell(ws.cell(row=wsRow,column=wsCol+ 9))

          if (entry.location.GetVal() == 128):
            self.flagCell(ws.cell(row=wsRow,column=wsCol+ 7))
            self.setCell(ws.cell(row=wsRow,column=wsCol+12),'L','G',entry.note.GetWSVal())
            self.flagCell(ws.cell(row=wsRow,column=wsCol+12))

          if (entry.product.GetVal() == 22):
            self.flagCell(ws.cell(row=wsRow,column=wsCol+ 9))
            self.setCell(ws.cell(row=wsRow,column=wsCol+12),'L','G',entry.note.GetWSVal())
            self.flagCell(ws.cell(row=wsRow,column=wsCol+12))

          if (entry.hours.GetVal() == None):
            self.flagCell(ws.cell(row=wsRow,column=wsCol+10))

          if (entry.workType.GetVal() == None):
            self.flagCell(ws.cell(row=wsRow,column=wsCol+11))

          # TODO: if hours == 0 or hours > 12 FLAG
          # TODO: Blue Strips
          # TODO: Check Cust vs Loc
          # TODO: Add Local/Remote
          # TODO: Add Local Text (AM-NE, AM-BA, etc)

          metrics.Update(fae,entry)

          # end of entry
          wsRow += 1

      #---------------------------------------------------------------
      # end of a week
      self.metrics[wsDate] = metrics
      wsRow += 2

    #-----------------------------------------------------------------
    # Write out raw metrics
    #-----------------------------------------------------------------
    #for i in range(1,self.week+1):
    #  wsDate = Calendar.week[i]
    #  metrics = self.metrics[wsDate]
    #  metrics.Log()

    wsRow = 2
    wsCol = 2

    ws_name = region + ' Metrics'
    ws = self.swb.get_sheet_by_name(ws_name)

    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 10

    i = 1
    while (i < len(TSSummary.colName)):
      self.setCell(ws.cell(row=i+3,column=wsCol+0),'L','G',TSSummary.colName[i])
      i += 1

    for fae in team.list:
      name = fae.fullname.GetVal()
      self.setCell(ws.cell(row=i+3,column=wsCol+0),'L','G','    ' + name)
      i += 1

    for i in range(1,self.week+1):
      wsDate = Calendar.week[i]
      metrics = self.metrics[wsDate]
      self.setCell(ws.cell(row= 2,column=wsCol+0+i),'C','G',str(wsDate))
      self.setCell(ws.cell(row= 3,column=wsCol+0+i),'C','G',str(i))
      self.setCell(ws.cell(row= 6,column=wsCol+0+i),'R','F',metrics.codes.kam.erc)
      self.setCell(ws.cell(row= 7,column=wsCol+0+i),'R','F',metrics.codes.kam.nok)
      self.setCell(ws.cell(row= 8,column=wsCol+0+i),'R','F',metrics.codes.kam.alu)
      self.setCell(ws.cell(row= 9,column=wsCol+0+i),'R','F',metrics.codes.kam.oth)
      self.setCell(ws.cell(row=11,column=wsCol+0+i),'R','F',metrics.codes.rka.qcm)
      self.setCell(ws.cell(row=12,column=wsCol+0+i),'R','F',metrics.codes.rka.att)
      self.setCell(ws.cell(row=13,column=wsCol+0+i),'R','F',metrics.codes.rka.spr)
      self.setCell(ws.cell(row=15,column=wsCol+0+i),'R','F',metrics.codes.car.att)
      self.setCell(ws.cell(row=16,column=wsCol+0+i),'R','F',metrics.codes.car.spr)
      self.setCell(ws.cell(row=17,column=wsCol+0+i),'R','F',metrics.codes.car.tmo)
      self.setCell(ws.cell(row=19,column=wsCol+0+i),'R','F',metrics.codes.smc.qcm)
      self.setCell(ws.cell(row=20,column=wsCol+0+i),'R','F',metrics.codes.smc.itl)
      self.setCell(ws.cell(row=21,column=wsCol+0+i),'R','F',metrics.codes.smc.prw)
      self.setCell(ws.cell(row=22,column=wsCol+0+i),'R','F',metrics.codes.smc.spd)
      self.setCell(ws.cell(row=24,column=wsCol+0+i),'R','F',metrics.codes.mod.qor)
      self.setCell(ws.cell(row=25,column=wsCol+0+i),'R','F',metrics.codes.mod.ter)
      self.setCell(ws.cell(row=27,column=wsCol+0+i),'R','F',metrics.codes.sem.air)
      self.setCell(ws.cell(row=28,column=wsCol+0+i),'R','F',metrics.codes.sem.van)
      self.setCell(ws.cell(row=30,column=wsCol+0+i),'R','F',metrics.codes.oth.oth)
      self.setCell(ws.cell(row=31,column=wsCol+0+i),'R','F',metrics.codes.oth.cob)
      self.setCell(ws.cell(row=32,column=wsCol+0+i),'R','F',metrics.codes.oth.ttt)
      self.setCell(ws.cell(row=34,column=wsCol+0+i),'R','F',metrics.codes.ovr.x4x)
      self.setCell(ws.cell(row=35,column=wsCol+0+i),'R','F',metrics.codes.ovr.x1x)
      self.setCell(ws.cell(row=37,column=wsCol+0+i),'R','F',metrics.team.dmr)
      self.setCell(ws.cell(row=38,column=wsCol+0+i),'R','F',metrics.team.mi)
      self.setCell(ws.cell(row=40,column=wsCol+0+i),'R','F',metrics.total.tot)
      self.setCell(ws.cell(row=41,column=wsCol+0+i),'R','F',metrics.total.prm)
      self.setCell(ws.cell(row=42,column=wsCol+0+i),'R','F',metrics.total.con)
      self.setCell(ws.cell(row=43,column=wsCol+0+i),'R','F',metrics.total.lbr)
      self.setCell(ws.cell(row=44,column=wsCol+0+i),'R','F',metrics.total.trv)
      self.setCell(ws.cell(row=45,column=wsCol+0+i),'R','F',metrics.total.sby)
      rowoffs = 0
      for fae in team.list:
        name = fae.fullname.GetVal()
        if (name in metrics.fae.dict):
          hours = metrics.fae.dict[name].hours
        else:
          hours = None
        self.setCell(ws.cell(row=47+rowoffs,column=wsCol+0+i),'R','F',hours)
        if (hours == None):
          weDate = wsDate + datetime.timedelta(days=4) 
          sDate = team.dict[name].startDate.GetVal()
          tDate = team.dict[name].endDate.GetVal()
          if (wsDate >= sDate and weDate <= tDate):
            self.flagCell(ws.cell(row=27+rowoffs,column=wsCol+0+i))

        rowoffs += 1

  #-----------------------------------------------------------------------
  def Sort(self,tsdata,week):
    self.sslist = []
    for i in range(1,week+1):
      list = []
      dict = {}
      wsDate = Calendar.week[i]
      dict = tsdata.weeks[wsDate]
      for key,value in dict.items():
        list.append(value)
      self.sslist.append(sorted(list))
 





