import os
import datetime
import logging
from openpyxl import load_workbook

#-----------------------------------------------------------------------
class FAEMember:
  def __init__(self,fname,lname,laborType,team,loc):
    self.fname     = fname
    self.lname     = lname
    self.fullname  = fname + ' ' + lname
    self.laborType = laborType
    self.team      = team
    self.loc       = loc

#-----------------------------------------------------------------------
class FAETeam:
  def __init__(self):
    self.members = {}
    self.members['Wray Odom'         ] = FAEMember('Wray'    ,'Odom'      ,'P','DMR','E')
    self.members['Pankaj Wadhwa'     ] = FAEMember('Pankaj'  ,'Wadhwa'    ,'C','DMR','E')
    self.members['Andy Cooper'       ] = FAEMember('Andy'    ,'Cooper'    ,'P','DMR','E')
    self.members['Kapil Bhardwaj'    ] = FAEMember('Kapil'   ,'Bhardwaj'  ,'P','DMR','E')
    self.members['Sohan D\'Souza'    ] = FAEMember('Sohan'   ,'D\'Souza'  ,'P','DMR','E')
    self.members['Jeremy Schroeder'  ] = FAEMember('Jeremy'  ,'Schroder'  ,'P','DMR','E')
    self.members['Emad Ramahi'       ] = FAEMember('Emad'    ,'Ramahi'    ,'P','DMR','E')
    self.members['Karun Dua'         ] = FAEMember('Karun'   ,'Dua'       ,'P','DMR','W')
    self.members['Suarabh Dhancholia'] = FAEMember('Suarabh' ,'Dhancholia','P','DMR','W')
    self.members['Jeff Smith'        ] = FAEMember('Jeff'    ,'Smith'     ,'P','DMR','W')
    self.members['Paul Khatkar'      ] = FAEMember('Paul'    ,'Khatkar'   ,'P','DMR','W')
    self.members['Joel Joseph'       ] = FAEMember('Joel'    ,'Joseph'    ,'P','DMR','W')
    self.members['Ashwini Bhagat'    ] = FAEMember('Ashwini' ,'Bhagat'    ,'P','DMR','W')
    self.members['Jim Morrison'      ] = FAEMember('Jim'     ,'Morrison'  ,'P','MI' ,'R')
    self.members['Paul Moser'        ] = FAEMember('Paul'    ,'Moser'     ,'P','MI' ,'R')
    self.members['Karl Hornung'      ] = FAEMember('Karl'    ,'Hornung'   ,'P','MI' ,'R')
    self.members['Jonathan Smith'    ] = FAEMember('Jonathan','Smith'     ,'P','MI' ,'R')
    self.members['Pouyan Rostam'     ] = FAEMember('Pouyan'  ,'Rostam'    ,'C','DMR','C')
    self.members['Karan Kalsi'       ] = FAEMember('Karan'   ,'Kalsi'     ,'C','DMR','C')
    self.members['Pallavi Chaturvedi'] = FAEMember('Pallavi' ,'Chaturvedi','C','DMR','C')

  def Append(self, fae):
    self.members.append(fae)

#-----------------------------------------------------------------------
class TSFile:
  def __init__(self,filename):
    self.fae   = None
    self.valid = False
    filenameinfo = os.path.basename(filename).split();
    if (len(filenameinfo) != 7):
      return

    year  = int(filenameinfo[6][0:4])
    if (year < 2016 or year > 2020): return
    month = int(filenameinfo[6][5:7])
    if (month < 1 or month > 12): return
    day   = int(filenameinfo[6][8:10])
    if (day < 1 or day > 31): return

    self.fname    = filenameinfo[2]
    self.lname    = filenameinfo[3]
    self.fullname = self.fname + ' ' + self.lname
    self.filename = filename
    self.weDate   = datetime.date(year, month, day)
    self.wsDate   = self.weDate - datetime.timedelta(days=6)
    self.valid    = True

    #check day of week; should be mon and sun

  #def __repr__(self):
  #  return '{}: {} {}'.format(self.__class__.__name__,self.wsDate,self.fullname)

  #def TSFileCompare(x,y):
  #  if (self.fae.team < other.fae.team): return other
  #  if (self.fae.team == other.fae.team):
  #    if (self.fae.loc < other.fae.loc):return other
  #  if (self.lname < other.lname): return other
  #  if (self.lname == other.lname):
  #    if (self.fname < other.fname): return other
  #  return self

  def __lt__(self,other): 
    #logging.debug(self.fae.team + ' ' + self.fae.loc + ' ' + self.fae.fullname + ' - ' + other.fae.team + ' ' + other.fae.loc + ' ' + other.fae.fullname)
    if (self.fae.team != other.fae.team):
      if (self.fae.team < other.fae.team):
        #logging.debug(self.fae.team + ' is LT')
        return True
      else:
        return False
    elif (self.fae.loc != other.fae.loc):
      if (self.fae.loc < other.fae.loc):
        #logging.debug(self.fae.loc + ' is LT')
        return True
      else:
        return False
    elif (self.fae.lname != other.fae.lname):
      if (self.fae.lname < other.fae.lname):
        #logging.debug(self.fae.lname + ' is LT')
        return True
      else:
        return False
    elif (self.fae.fname != other.fae.fname):
      if (self.fae.fname < other.fae.fname):
        #logging.debug(self.fae.fname + ' is LT')
        return True
      else:
        return False
    logging.error('SHOULD NOT BE HERE')
    return False
    #if (self.fae.team  < other.fae.team):
    #  logging.debug(self.fae.team + ' is LT') 
    #  return True
    #elif (self.fae.loc < other.fae.loc):
    #  logging.debug(self.fae.loc + ' is LT')
    #  return True
    #elif (self.fae.lname < other.fae.lname):
    #  logging.debug(self.fae.lname + ' is LT')
    #  return True
    #elif (self.fae.fname < other.fae.fname):
    #  logging.debug(self.fae.fname + ' is LT') 
    #  return True
  def __le__(self,other): 
    return True
  def __gt__(self,other): 
    return True
  def __ge__(self,other): 
    return True
  def __eq__(self,other): 
    return True
  def __ne__(self,other): 
    return True

  def AddFAE(self,fae):
    self.fae = fae

  def IsValid(self):
    return (self.valid == True and self.fae != None)

  def Log(self):
    logging.debug(str(self.wsDate) + ' ' + self.fname + ' ' + self.lname)

#-----------------------------------------------------------------------
class TSData:
  def __init__(self):
    self.weeks = {}

  def AddFile(self,tsfile):
    date = tsfile.wsDate
    if (date not in self.weeks):
      self.weeks[date] = {}
    name = tsfile.fullname
    if (name not in self.weeks[date]):
      self.weeks[date][name] = tsfile
    else:
      logging.error('Timesheet already exists for ' + str(date) + ' ' + name)

  def Validate(self, faes):
    for date in self.weeks:
      cnt = 0
      for name in faes.members:
        if (name not in self.weeks[date]):
          logging.error('Missing Timesheet for ' + str(date) + ' ' + name)
        else:
          cnt += 1
      logging.debug('Week ' + str(date) + ' Cnt ' + str(cnt))

  def Log(self):
    for week in self.weeks:
      logging.debug('Week' + ' ' + str(week))
      for name in self.weeks[week]:
        logging.debug('Name' + '            ' + name)

#-----------------------------------------------------------------------
class Week:
  def __init__(self):
    self.date = None
    self.tsinfo = []

#-----------------------------------------------------------------------
#class TSInfo:
#  def __init__(self,fnameinfo):
#    self.fname    = fnameinfo.fname
#    self.lname    = fnameinfo.lname
#    self.fullname = fnameinfo.fullname
#    self.filename = fnameinfo.filename
#    self.weDate   = fnameinfo.weDate
#    self.wsDate   = fnameinfo.wsDate
#    self.valid    = True
#
#  def Log(self):
#    logging.debug(str(self.wsDate) + ' ' + self.fname + ' ' + self.lname)


#-----------------------------------------------------------------------
class RegionReport:
  def __init__(self):
    self.week = {}

#-----------------------------------------------------------------------
class TSCalendar:
  def __init__(self,year):
    self.week = {}
    d = datetime.date(year,1,1)
    while (True):
      #logging.debug(d.weekday())
      if (d.weekday() == 0):
        break;
      else:
        d = d + datetime.timedelta(days=1)
    #logging.debug(d)
    i = 1
    while(True):
      self.week[i] = d
      #logging.debug(str(i) + ' ' + str(self.week[i]))
      d = d + datetime.timedelta(days = 7)
      if (d.year > year):
        break;
      i += 1
   
  def Log(self):
    cnt = len(self.week)
    for i in range(1,cnt+1):
      logging.debug(str(i) + ' ' + str(self.week[i])) 
 

#-----------------------------------------------------------------------
# Class Timesheet Entry
#-----------------------------------------------------------------------
class TSEntry:
  def __init__(self,date,partA,partB,partC,partD,hours,ltd,notes):
    self.date  = date
    self.partA = partA
    self.partB = partB
    self.partC = partC
    self.partD = partD
    self.hours = hours
    self.ltd   = ltd
    self.notes = notes

  def Log(self):
    date  = self.date
    partA = self.partA.ljust(50)
    partB = self.partB.ljust(60)
    partC = self.partC.ljust(45)
    partD = self.partD.ljust(35)
    hours = str(self.hours).ljust(5)
    ltd   = self.ltd.ljust(10)
    logging.debug(str(date) + '|' + partA + '|' + partB + '|' + partC + '|' + partD + '|' + hours + '|' + ltd)

#-----------------------------------------------------------------------
# Class Timesheet
#-----------------------------------------------------------------------
class Timesheet:
  def __init__(self, ssdata, wsDate):
    self.entries = []
    self.ssdata  = ssdata
    self.wsDate  = wsDate

  #---------------------------------------------------------------------
  def calcDate(self,day):
    day = day.lower().strip()
    day = day[0:3]
    day = { 'mon':0,'tue':1,'wed':2,'thu':3,'fri':4,'sat':5,'sun':6 }.get(day,7)
    if (day == 7):
      logging.error('Invalid Day')
    date = self.wsDate + datetime.timedelta(days = day)
    return date

  #---------------------------------------------------------------------
  def Clean(self):
    for i,entry in enumerate(self.entries):
      entry.Log()

  #---------------------------------------------------------------------
  def ReadFile(self):
    path = self.ssdata.filename
    logging.debug(path)
    wb = load_workbook(path)
    sheet_names = wb.get_sheet_names();
    if (sheet_names[0] != 'Timesheet'):
      logging.error(path + ' is not a valid Timesheet spreadsheet')

    ws = wb.get_sheet_by_name('Timesheet')

    wsRow     = 7
    wsCol     = 4
    sundayFlg = False
    sundayCnt = 0;
    blankFlg  = False
    curDate   = None

    #syncCell = ws.cell(row=6,column=4)
    #logging.debug(syncCell)
    if (ws.cell(row=6,column=4).value == 'Customer - Part A'): 
      wsCol = 3
    elif (ws.cell(row=6,column=5).value == 'Customer - Part A'):
      wsCol = 4
    else:
      logging.error('Couldn\'t sync to Timesheet spreadsheet')
      return

    while True:
      day = ws.cell(row=wsRow,column=wsCol+0).value
      if (day == None):
        day = ''
        if (curDate == None):
          logging.error('SHOULDNT BE HERE')
      else:
        curDate = self.calcDate(day)
 
      partA  = ws.cell(row=wsRow,column=wsCol+1).value
      if (partA == None):
        partA = ''
        blankFlg = True

      partB = ws.cell(row=wsRow,column=wsCol+2).value
      if (partB == None):
        partB = ''

      partC = ws.cell(row=wsRow,column=wsCol+3).value
      if (partC == None):
        partC = ''

      partD = ws.cell(row=wsRow,column=wsCol+4).value
      if (partD == None):
        partD = ''

      hours = ws.cell(row=wsRow,column=wsCol+9).value
      #logging.debug('hours ' + str(hours))
      if (hours == None):
        hours = ''
      else:
        try:
          hours = float(hours)
        except:
          hours = ''
      #logging.debug('hours ' + str(hours))
      #logging.debug('hours ' + str(hours))

      ltd = ws.cell(row=wsRow,column=wsCol+10).value
      if (ltd == None):
        ltd = ''

      notes = ws.cell(row=wsRow,column=wsCol+11).value
      if (notes == None):
        notes = ''
 
      # TODO: if blankFlg == False, then make sure everything else is blank

      if (not blankFlg):
        #if (day == ''):
        #  logging.error('Day is blank')
        #logging.debug(curDay + '|' + partA + '|' + partB + '|' + partC + '|' + '|')
        self.entries.append(TSEntry(curDate,partA,partB,partC,partD,hours,ltd,notes))

      if (day.startswith('Sunday')):
        sundayFlg = True
      if (sundayFlg):
        if (blankFlg):
          sundayCnt += 1
        else:
          sundayCnt = 0;
      if (sundayCnt > 5):
        break

      wsRow += 1
      blankFlg = False

    #logging.debug('Done ReadFile')

  def Log(self):
    for i,entry in enumerate(self.entries):
      fullname = self.ssdata.fae.fullname.ljust(25)
      date  = self.entries[i].date
      partA = self.entries[i].partA.ljust(50)
      partB = self.entries[i].partB.ljust(60)
      partC = self.entries[i].partC.ljust(45)
      partD = self.entries[i].partD.ljust(35)
      hours = str(self.entries[i].hours).ljust(5)
      ltd   = self.entries[i].ltd.ljust(10)
      logging.debug(fullname + '|' + str(date) + '|' + partA + '|' + partB + '|' + partC + '|' + partD + '|' + hours + '|' + ltd)

#-----------------------------------------------------------------------
# Class TSSummary
#-----------------------------------------------------------------------
class TSSummary:
  def __init__(self):
    self.sslist = []
    self.tsdict = {}
    self.week = None
    self.year = None

  #---------------------------------------------------------------------
  def Process(self,tsdata,team,year,week):
    self.year = year
    self.week = week

    cal = TSCalendar(year)
    self.Sort(tsdata,cal,week)

    for i,week in enumerate(self.sslist):
      wsDate = cal.week[i+1]
      logging.debug('***' + str(wsDate) + '***')
      self.tsdict[wsDate] = []
      for j,ssdata in enumerate(week):
        ts = Timesheet(ssdata, wsDate)
        ts.ReadFile()
        ts.Clean()
        self.tsdict[wsDate].append(ts)
      logging.debug('***' + str(wsDate) + '***')

    #for i in range(0,self.week-1):
    #  wsDate = cal.week[i+1]
    #  for j,ts in enumerate(self.tsdict[wsDate]):
    #    self.tsdict[wsDate][j].Log()

  #-----------------------------------------------------------------------
  def Sort(self,tsdata,cal,week):
    self.sslist = []
    for i in range(1,week+1):
      list = []
      dict = {}
      wsDate = cal.week[i]
      logging.debug(str(i) + ' ' + str(wsDate))
      dict = tsdata.weeks[wsDate]
      for key,value in dict.items():
        list.append(value)
      self.sslist.append(sorted(list))
      #list.sort(key=list.Compare)
      #for j in list:
      #  logging.debug(j.fae.team + ' ' + j.fae.loc + ' ' + j.fae.fullname)
 
    for i,week in enumerate(self.sslist):
      logging.debug(cal.week[i+1])
      for j,item in enumerate(week):
        logging.debug(item.fae.team + ' ' + item.fae.loc + ' ' + item.fae.fullname)
    #for i in range(len(self.sslist)):
    #  logging.debug(cal.week[k])
    #  for j in range(len(self.sslist[i])):
    #    logging.debug(j.fae.team + ' ' + j.fae.loc + ' ' + j.fae.fullname)
    #  k += 1
        

    logging.debug('sorted list')






