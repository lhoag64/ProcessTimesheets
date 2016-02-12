import os
import datetime
import logging
from openpyxl import load_workbook

#-----------------------------------------------------------------------
class FAEMember:
  def __init__(self,fname,lname,laborType,team,loc):
    self.fname     = fname
    self.lname     = lname
    self.fullname  = fname + ' ' + lname
    self.laborType = laborType
    self.team      = team
    self.loc       = loc

#-----------------------------------------------------------------------
class FAETeam:
  def __init__(self):
    self.members = {}
    self.members['Wray Odom'         ] = FAEMember('Wray'    ,'Odom'      ,'P','DMR','E')
    self.members['Pankaj Wadhwa'     ] = FAEMember('Pankaj'  ,'Wadhwa'    ,'C','DMR','E')
    self.members['Andy Cooper'       ] = FAEMember('Andy'    ,'Cooper'    ,'P','DMR','E')
    self.members['Kapil Bhardwaj'    ] = FAEMember('Kapil'   ,'Bhardwaj'  ,'P','DMR','E')
    self.members['Sohan D\'Souza'    ] = FAEMember('Sohan'   ,'D\'Souza'  ,'P','DMR','E')
    self.members['Jeremy Schroeder'  ] = FAEMember('Jeremy'  ,'Schroder'  ,'P','DMR','E')
    self.members['Emad Ramahi'       ] = FAEMember('Emad'    ,'Ramahi'    ,'P','DMR','E')
    self.members['Karun Dua'         ] = FAEMember('Karun'   ,'Dua'       ,'P','DMR','W')
    self.members['Suarabh Dhancholia'] = FAEMember('Suarabh' ,'Dhancholia','P','DMR','W')
    self.members['Jeff Smith'        ] = FAEMember('Jeff'    ,'Smith'     ,'P','DMR','W')
    self.members['Paul Khatkar'      ] = FAEMember('Paul'    ,'Khatkar'   ,'P','DMR','W')
    self.members['Joel Joseph'       ] = FAEMember('Joel'    ,'Joseph'    ,'P','DMR','W')
    self.members['Ashwini Bhagat'    ] = FAEMember('Ashwini' ,'Bhagat'    ,'P','DMR','W')
    self.members['Jim Morrison'      ] = FAEMember('Jim'     ,'Morrison'  ,'P','MI' ,'R')
    self.members['Paul Moser'        ] = FAEMember('Paul'    ,'Moser'     ,'P','MI' ,'R')
    self.members['Karl Hornung'      ] = FAEMember('Karl'    ,'Hornung'   ,'P','MI' ,'R')
    self.members['Jonathan Smith'    ] = FAEMember('Jonathan','Smith'     ,'P','MI' ,'R')
    self.members['Pouyan Rostam'     ] = FAEMember('Pouyan'  ,'Rostam'    ,'C','DMR','C')
    self.members['Karan Kalsi'       ] = FAEMember('Karan'   ,'Kalsi'     ,'C','DMR','C')
    self.members['Pallavi Chaturvedi'] = FAEMember('Pallavi' ,'Chaturvedi','C','DMR','C')

  def Append(self, fae):
    self.members.append(fae)

#-----------------------------------------------------------------------
class FLFile:
  def __init__(self,filename):
    self.fae   = None
    self.valid = False
    filenameinfo = os.path.basename(filename).split();
    if (len(filenameinfo) != 7):
      return

    year  = int(filenameinfo[6][0:4])
    if (year < 2016 or year > 2020): return
    month = int(filenameinfo[6][5:7])
    if (month < 1 or month > 12): return
    day   = int(filenameinfo[6][8:10])
    if (day < 1 or day > 31): return

    self.fname    = filenameinfo[2]
    self.lname    = filenameinfo[3]
    self.fullname = self.fname + ' ' + self.lname
    self.filename = filename
    self.weDate   = datetime.date(year, month, day)
    self.wsDate   = self.weDate - datetime.timedelta(days=6)
    self.valid    = True

  def __lt__(self,other): 
    if (self.fae.team != other.fae.team):
      if (self.fae.team < other.fae.team):
        return True
      else:
        return False
    elif (self.fae.loc != other.fae.loc):
      if (self.fae.loc < other.fae.loc):
        return True
      else:
        return False
    elif (self.fae.lname != other.fae.lname):
      if (self.fae.lname < other.fae.lname):
        return True
      else:
        return False
    elif (self.fae.fname != other.fae.fname):
      if (self.fae.fname < other.fae.fname):
        return True
      else:
        return False
    logging.error('SHOULD NOT BE HERE')
    return False
  def __le__(self,other): 
    return True
  def __gt__(self,other): 
    return True
  def __ge__(self,other): 
    return True
  def __eq__(self,other): 
    return True
  def __ne__(self,other): 
    return True

  def AddFAE(self,fae):
    self.fae = fae

  def IsValid(self):
    return (self.valid == True and self.fae != None)

  def Log(self):
    logging.debug(str(self.wsDate) + ' ' + self.fname + ' ' + self.lname)

#-----------------------------------------------------------------------
class FLData:
  def __init__(self):
    self.weeks = {}

  def AddFile(self,tsfile):
    date = tsfile.wsDate
    if (date not in self.weeks):
      self.weeks[date] = {}
    name = tsfile.fullname
    if (name not in self.weeks[date]):
      self.weeks[date][name] = tsfile
    else:
      logging.error('Timesheet already exists for ' + str(date) + ' ' + name)

  def Validate(self, faes):
    for date in self.weeks:
      cnt = 0
      for name in faes.members:
        if (name not in self.weeks[date]):
          logging.error('Missing Timesheet for ' + str(date) + ' ' + name)
        else:
          cnt += 1
      logging.debug('Week ' + str(date) + ' Cnt ' + str(cnt))

  def Log(self):
    for week in self.weeks:
      logging.debug('Week' + ' ' + str(week))
      for name in self.weeks[week]:
        logging.debug('Name' + '            ' + name)

#-----------------------------------------------------------------------
#class Week:
#  def __init__(self):
#    self.date = None
#    self.tsinfo = []

#-----------------------------------------------------------------------
#class RegionReport:
#  def __init__(self):
#    self.week = {}

# TS == TimeSheet
# FN == FileName
# SS == SpreadSheet
# Summary ? DB == DashBoard ?
# TSTemplate

#-----------------------------------------------------------------------
class Calendar:
  def __init__(self,year):
    self.week = {}
    d = datetime.date(year,1,1)
    while (True):
      if (d.weekday() == 0):
        break;
      else:
        d = d + datetime.timedelta(days=1)
    i = 1
    while(True):
      self.week[i] = d
      d = d + datetime.timedelta(days = 7)
      if (d.year > year):
        break;
      i += 1
   
  def Log(self):
    cnt = len(self.week)
    for i in range(1,cnt+1):
      logging.debug(str(i) + ' ' + str(self.week[i])) 
 
#-----------------------------------------------------------------------
class TSDate:
  def __init__(self, date):
    self.asStr = str(date)

  def __str__(self): return self.asStr

#-----------------------------------------------------------------------
class TSCode:
  def __init__(self, partA):
    self.asStr = partA

  def __str__(self): return self.asStr

  def Parse(self): pass
  def Clean (self): pass
  def Log(self): pass

#-----------------------------------------------------------------------
class TSLocation:
  def __init__(self,partB):
    self.asStr = partB

  def __str__(self): return self.asStr

#-----------------------------------------------------------------------
class TSActivity:
  def __init__(self,partC):
    self.asStr = partC

  def __str__(self): return self.asStr

#-----------------------------------------------------------------------
class TSProduct:
  def __init__(self,partD):
    self.asStr = partD

  def __str__(self): return self.asStr

#-----------------------------------------------------------------------
class TSHours:
  def __init__(self,hours):
    self.asStr = str(hours)
    self.asNum = hours

  def __str__(self): return self.asStr

#-----------------------------------------------------------------------
class TSWorkType:
  def __init__(self,workType):
    self.asStr = workType

  def __str__(self): return self.asStr

#-----------------------------------------------------------------------
class TSNote:
  def __init__(self,note):
    self.asStr = note

  def __str__(self): return self.asStr

#-----------------------------------------------------------------------
# Class Timesheet Entry
#-----------------------------------------------------------------------
class TSEntry:
  def __init__(self,date,partA,partB,partC,partD,hours,workType,note):
    self.date     = TSDate(date)
    self.code     = TSCode(partA)
    self.location = TSLocation(partB)
    self.activity = TSActivity(partC)
    self.product  = TSProduct(partD)
    self.hours    = TSHours(hours)
    self.workType = TSWorkType(workType)
    self.note     = TSNote(note)

  def Log(self):
    date  = str(self.date)
    partA = str(self.code).ljust(50)     # Customer/Code
    partB = str(self.location).ljust(60) # Location
    partC = str(self.activity).ljust(45) # Activity
    partD = str(self.product).ljust(35)  # Product
    hours = str(self.hours).ljust(5)
    ltd   = str(self.workType).ljust(10)
    logging.debug(str(date) + '|' + partA + '|' + partB + '|' + partC + '|' + partD + '|' + hours + '|' + ltd)

#-----------------------------------------------------------------------
# Class Timesheet
#-----------------------------------------------------------------------
class Timesheet:
  def __init__(self, ssdata, wsDate):
    self.entries = []
    self.ssdata  = ssdata
    self.wsDate  = wsDate

  #---------------------------------------------------------------------
  def calcDate(self,day):
    day = day.lower().strip()
    day = day[0:3]
    day = { 'mon':0,'tue':1,'wed':2,'thu':3,'fri':4,'sat':5,'sun':6 }.get(day,7)
    if (day == 7):
      logging.error('Invalid Day')
    date = self.wsDate + datetime.timedelta(days = day)
    return date

  #---------------------------------------------------------------------
  def Clean(self):
    for i,entry in enumerate(self.entries):
      entry.Log()

  #---------------------------------------------------------------------
  def ReadFile(self):
    path = self.ssdata.filename
    logging.debug(path)
    wb = load_workbook(path)
    sheet_names = wb.get_sheet_names();
    if (sheet_names[0] != 'Timesheet'):
      logging.error(path + ' is not a valid Timesheet spreadsheet')

    ws = wb.get_sheet_by_name('Timesheet')

    wsRow     = 7
    wsCol     = 4
    sundayFlg = False
    sundayCnt = 0;
    blankFlg  = False
    curDate   = None

    if (ws.cell(row=6,column=4).value == 'Customer - Part A'): 
      wsCol = 3
    elif (ws.cell(row=6,column=5).value == 'Customer - Part A'):
      wsCol = 4
    else:
      logging.error('Couldn\'t sync to Timesheet spreadsheet')
      return

    while True:
      day = ws.cell(row=wsRow,column=wsCol+0).value
      if (day == None):
        day = ''
        if (curDate == None):
          logging.error('SHOULDNT BE HERE')
      else:
        curDate = self.calcDate(day)
 
      partA  = ws.cell(row=wsRow,column=wsCol+1).value
      if (partA == None):
        partA = ''
        blankFlg = True

      partB = ws.cell(row=wsRow,column=wsCol+2).value
      if (partB == None):
        partB = ''

      partC = ws.cell(row=wsRow,column=wsCol+3).value
      if (partC == None):
        partC = ''

      partD = ws.cell(row=wsRow,column=wsCol+4).value
      if (partD == None):
        partD = ''

      hours = ws.cell(row=wsRow,column=wsCol+9).value
      if (hours == None):
        hours = ''
      else:
        try:
          hours = float(hours)
        except:
          hours = ''

      ltd = ws.cell(row=wsRow,column=wsCol+10).value
      if (ltd == None):
        ltd = ''

      notes = ws.cell(row=wsRow,column=wsCol+11).value
      if (notes == None):
        notes = ''
 
      # TODO: if blankFlg == False, then make sure everything else is blank

      if (not blankFlg):
        self.entries.append(TSEntry(curDate,partA,partB,partC,partD,hours,ltd,notes))

      if (day.startswith('Sunday')):
        sundayFlg = True
      if (sundayFlg):
        if (blankFlg):
          sundayCnt += 1
        else:
          sundayCnt = 0;
      if (sundayCnt > 5):
        break

      wsRow += 1
      blankFlg = False

  #---------------------------------------------------------------------
  def Log(self):
    for i,entry in enumerate(self.entries):
      fullname = self.ssdata.fae.fullname.ljust(25)
      date  = str(self.entries[i].date)
      partA = str(self.entries[i].code).ljust(50)
      partB = str(self.entries[i].location).ljust(60)
      partC = str(self.entries[i].activity).ljust(45)
      partD = str(self.entries[i].product).ljust(35)
      hours = str(self.entries[i].hours).ljust(5)
      ltd   = str(self.entries[i].worktype).ljust(10)
      logging.debug(fullname + '|' + str(date) + '|' + partA + '|' + partB + '|' + partC + '|' + partD + '|' + hours + '|' + ltd)

#-----------------------------------------------------------------------
# Class TSSummary
#-----------------------------------------------------------------------
class TSSummary:
  def __init__(self):
    self.sslist = []
    self.tsdict = {}
    self.week = None
    self.year = None

  #---------------------------------------------------------------------
  def Process(self,tsdata,team,year,week):
    self.year = year
    self.week = week

    cal = Calendar(year)
    self.Sort(tsdata,cal,week)

    for i,week in enumerate(self.sslist):
      wsDate = cal.week[i+1]
      logging.debug('***' + str(wsDate) + '***')
      self.tsdict[wsDate] = []
      for j,ssdata in enumerate(week):
        ts = Timesheet(ssdata, wsDate)
        ts.ReadFile()
        ts.Clean()
        self.tsdict[wsDate].append(ts)
      logging.debug('***' + str(wsDate) + '***')

  #-----------------------------------------------------------------------
  def Sort(self,tsdata,cal,week):
    self.sslist = []
    for i in range(1,week+1):
      list = []
      dict = {}
      wsDate = cal.week[i]
      logging.debug(str(i) + ' ' + str(wsDate))
      dict = tsdata.weeks[wsDate]
      for key,value in dict.items():
        list.append(value)
      self.sslist.append(sorted(list))
 
    for i,week in enumerate(self.sslist):
      logging.debug(cal.week[i+1])
      for j,item in enumerate(week):
        logging.debug(item.fae.team + ' ' + item.fae.loc + ' ' + item.fae.fullname)
        
    logging.debug('sorted list')






