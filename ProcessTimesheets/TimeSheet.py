#import os
import datetime
import logging
from   openpyxl import load_workbook
from   FAE      import FAETeam
 
#-----------------------------------------------------------------------
class TSDate:
  def __init__(self, date):
    self.asStr = str(date)

  def GetWSVal(self):
    if (len(self.asStr) == 0):
      return None
    else:
      return self.asStr

  def __str__(self): return self.asStr

#-----------------------------------------------------------------------
# Class encapsulating Work Codes (or customers)
#-----------------------------------------------------------------------
class TSCode:

  #---------------------------------------------------------------------
  # Static Variable cDict which holds valid codes, initialized by Init
  #---------------------------------------------------------------------
  cDict = None

  #---------------------------------------------------------------------
  # Initialize cDict
  #---------------------------------------------------------------------
  def Init(list):
    TSCode.cDict = {}
    for i,item in enumerate(list):
      tup = TSCode.Convert(item)
      TSCode.cDict.update({tup[0]:tup[1]})

  def Convert(text):
      tup = text.rpartition('-')
      desc = tup[0].strip()
      code = tup[2].strip()
      return (code,desc)

  #---------------------------------------------------------------------
  # Initialize an instance of TSCode
  #---------------------------------------------------------------------
  def __init__(self, partA):
    tup = TSCode.Convert(partA)
    if (tup[0] not in TSCode.cDict):
      raise Exception
    self.asStr = str(tup[0])
    if (len(self.asStr) == 3): 
      self.isCompany  = True
      self.isOverhead = False
    elif (len(self.asStr) == 5):
      self.isOverhead = True
      self.isCompany  = False
    else:
      self.isOverhead = None
      self.isCompany  = None

  def GetWSVal(self):
    if (len(self.asStr) == 0):
      return None
    else:
      return self.asStr

  def __str__(self):
    if (self.asStr != None): return self.asStr
    else: return ''

  def GetVal(self): return self.asStr

  def GetDesc(self):
    return TSCode.cDict[self.asStr][0:5]

#-----------------------------------------------------------------------
class TSLocation:
  #---------------------------------------------------------------------
  # Static Variable lDict which holds valid locations, initialized by Init
  #---------------------------------------------------------------------
  lDict = None

  #---------------------------------------------------------------------
  # Initialize lDict
  #---------------------------------------------------------------------
  def Init(list):
    TSLocation.lDict = {}
    for i,item in enumerate(list):
      tup = TSLocation.Convert(item)
      TSLocation.lDict.update({tup[0]:tup[1]})

  def Convert(text):
      tup = text.rpartition('-')
      desc = tup[0].strip()
      code = tup[2].strip()
      return (code,desc)

  #---------------------------------------------------------------------
  # Initialize an instance of TSLocation
  #---------------------------------------------------------------------
  def __init__(self,partB):
    if (len(partB) > 0):
      tup = TSLocation.Convert(partB)
      if (tup[0] not in TSLocation.lDict):
        raise Exception
      self.asStr = str(tup[0])
    else:
      self.asStr = None

  def GetWSVal(self):
    if (self.asStr != None):
      if (len(self.asStr) == 0):
        return None
      else:
        return int(self.asStr)
    else: return None

  def GetVal(self) : return self.asStr

  def __str__(self):
    if (self.asStr != None): return self.asStr
    else: return ''

#-----------------------------------------------------------------------
class TSActivity:
  #---------------------------------------------------------------------
  # Static Variable aDict which holds valid locations, initialized by Init
  #---------------------------------------------------------------------
  aDict = None

  #---------------------------------------------------------------------
  # Initialize aDict
  #---------------------------------------------------------------------
  def Init(list):
    TSActivity.aDict = {}
    for i,item in enumerate(list):
      tup = TSActivity.Convert(item)
      TSActivity.aDict.update({tup[0]:tup[1]})

  def Convert(text):
      tup = text.rpartition('-')
      desc = tup[0].strip()
      code = tup[2].strip()
      #logging.debug(code + '|' + desc + '|')
      return (code,desc)

  #---------------------------------------------------------------------
  # Initialize an instance of TSActivity
  #---------------------------------------------------------------------
  def __init__(self,partC):
    if (len(partC) > 0):
      tup = TSActivity.Convert(partC)
      if (tup[0] not in TSActivity.aDict):
        raise Exception
      self.asStr = str(tup[0])
    else:
      self.asStr = None

  def GetWSVal(self):
    if (self.asStr != None):
      if (len(self.asStr) == 0):
        return None
      else:
        return int(self.asStr)
    else: return None

  def GetVal(self) : return self.asStr

  def __str__(self):
    if (self.asStr != None): return self.asStr
    else: return ''

#-----------------------------------------------------------------------
class TSProduct:
  #---------------------------------------------------------------------
  # Static Variable pDict which holds valid locations, initialized by Init
  #---------------------------------------------------------------------
  pDict = None

  #---------------------------------------------------------------------
  # Initialize pDict
  #---------------------------------------------------------------------
  def Init(list):
    TSProduct.pDict = {}
    for i,item in enumerate(list):
      tup = TSProduct.Convert(item)
      TSProduct.pDict.update({tup[0]:tup[1]})

  def Convert(text):
      tup = text.rpartition('-')
      desc = tup[0].strip()
      code = tup[2].strip()
      return (code,desc)

  #---------------------------------------------------------------------
  # Initialize an instance of TSProduct
  #---------------------------------------------------------------------
  def __init__(self,partD):
    if (len(partD) > 0):
      tup = TSProduct.Convert(partD)
      if (tup[0] not in TSProduct.pDict):
        raise Exception
      self.asStr = str(tup[0])
    else:
      self.asStr = None

  def GetWSVal(self):
    if (self.asStr != None):
      if (len(self.asStr) == 0):
        return None
      else:
        return int(self.asStr)
    else: return None

  def GetVal(self) : return self.asStr

  def __str__(self):
    if (self.asStr != None): return self.asStr
    else: return ''

#-----------------------------------------------------------------------
class TSHours:
  def __init__(self,hours):
    self.asStr = str(hours)
    if (len(self.asStr) == 0):
      self.asNum = 0.0
    else:
      self.asNum = hours

  def GetWSVal(self):
    if (len(self.asStr) == 0):
      return None
    else:
      return self.asNum

  def GetVal(self) : return self.asNum

  def __str__(self): 
    if (self.asStr != None): return self.asStr
    else: return ''
  def __float__(self): return self.asNum

#-----------------------------------------------------------------------
class TSWorkType:
  wtSet = set(['Labour','Travel','Standby'])

  def __init__(self,workType):
    if (len(workType)):
      if (workType in TSWorkType.wtSet):
        self.asStr = workType
      else:
        self.asStr = None
    else:
      self.asStr = None

  def GetWSVal(self):
    if (self.asStr != None):
      if (len(self.asStr) == 0):
        return None
      else:
        return self.asStr
    else: return None

  def GetVal(self) : return self.asStr

  def __str__(self):
    if (self.asStr != None): return self.asStr
    else: return ''

#-----------------------------------------------------------------------
class TSNote:
  def __init__(self,note):
    if (note != None):
      note = str(note)
      if (len(note) == 0):
        self.asStr = ''
      else:
        self.asStr = note
    else:
      self.asStr = None

  def GetWSVal(self):
    if (len(self.asStr) == 0):
      return None
    else:
      return self.asStr[0:48]

  def __str__(self):
    if (self.asStr != None): return self.asStr
    else: return ''

#-----------------------------------------------------------------------
# Class Timesheet Entry
#-----------------------------------------------------------------------
class TSEntry:
  def __init__(self,date,partA,partB,partC,partD,hours,workType,note):
    self.date     = TSDate(date)
    self.code     = TSCode(partA)
    self.location = TSLocation(partB)
    self.activity = TSActivity(partC)
    self.product  = TSProduct(partD)
    self.hours    = TSHours(hours)
    self.workType = TSWorkType(workType)
    self.note     = TSNote(note)

  def Log(self):
    date  = str(self.date)
    partA = str(self.code).ljust(5)
    partB = str(self.location).ljust(3)  # Location
    partC = str(self.activity).ljust(2)  # Activity
    partD = str(self.product).ljust(2)   # Product
    hours = str(self.hours).ljust(5)
    ltd   = str(self.workType).ljust(10)
    #logging.debug(str(date) + '|' + partA + '|' + partB + '|' + partC + '|' + partD + '|' + hours + '|' + ltd)

#-----------------------------------------------------------------------
# Class Timesheet
#-----------------------------------------------------------------------
class Timesheet:
  def __init__(self,ssdata,wsDate):
    self.entries = []
    self.ssdata  = ssdata
    self.wsDate  = wsDate

  #---------------------------------------------------------------------
  def calcDate(self,day):
    day = day.lower().strip()
    day = day[0:3]
    day = { 'mon':0,'tue':1,'wed':2,'thu':3,'fri':4,'sat':5,'sun':6 }.get(day,7)
    if (day == 7):
      raise Exception
    date = self.wsDate + datetime.timedelta(days = day)
    return date

  #---------------------------------------------------------------------
  def Clean(self):
    for i,entry in enumerate(self.entries):
      entry.Log()

  #---------------------------------------------------------------------
  def ReadFile(self):
    path = self.ssdata.filename
    wb = load_workbook(path)
    sheet_names = wb.get_sheet_names();
    if (sheet_names[0] != 'Timesheet'):
      logging.error(path + ' is not a valid Timesheet spreadsheet')

    #logging.debug('Reading ' + path)

    ws = wb.get_sheet_by_name('Timesheet')

    wsRow     = 7
    wsCol     = 4
    sundayFlg = False
    sundayCnt = 0;
    blankFlg  = False
    curDate   = None

    if (ws.cell(row=6,column=4).value == 'Customer - Part A'): 
      wsCol = 3
    elif (ws.cell(row=6,column=5).value == 'Customer - Part A'):
      wsCol = 4
    else:
      logging.error('Couldn\'t sync to Timesheet spreadsheet')
      return

    # Try to find name
    found = False
    currow = 3
    curcol = 1
    while (curcol < 10):
      cellValue = str(ws.cell(row=currow,column=curcol).value)
      cellValue = cellValue.strip()
      if (cellValue == 'Name:'):
        found = True
        break
      curcol += 1
    if (found == True):
      #nameFromFile = str(ws.cell(row=currow,column=curcol+1).value).strip()
      nameFromFile = str(ws.cell(row=currow,column=curcol+2).value).strip()
      #nameFromFile = str(ws.cell(row=currow,column=curcol+3).value).strip()
    else:
      nameFromFile = ''

    # Try to find date
    found = False
    currow = 3
    curcol = 1
    while (curcol < 10):
      cellValue = str(ws.cell(row=currow,column=curcol).value)
      cellValue = cellValue.strip()
      if (cellValue[0:4] == 'Week'):
        found = True
        break
      curcol += 1
    if (found == True):
      dateFromFile = str(ws.cell(row=currow,column=curcol+1).value).strip()
      dateFromFile = dateFromFile[0:10]
    else:
      dateFromFile = ''

    nameFromFile = nameFromFile.ljust(15)
    dateFromFile = dateFromFile.ljust(15)

    logging.debug(nameFromFile + '|' + dateFromFile + '|Reading ' + path)

    while True:
      day = ws.cell(row=wsRow,column=wsCol+0).value
      if (day == None):
        day = ''
        if (curDate == None):
          logging.error('SHOULDNT BE HERE')
      else:
        curDate = self.calcDate(day)
 
      partA  = ws.cell(row=wsRow,column=wsCol+1).value
      if (partA == None):
        partA = ''
        blankFlg = True

      partB = ws.cell(row=wsRow,column=wsCol+2).value
      if (partB == None):
        partB = ''

      partC = ws.cell(row=wsRow,column=wsCol+3).value
      if (partC == None):
        partC = ''

      partD = ws.cell(row=wsRow,column=wsCol+4).value
      if (partD == None):
        partD = ''

      hours = ws.cell(row=wsRow,column=wsCol+9).value
      if (hours == None):
        hours = ''
      else:
        try:
          hours = float(hours)
        except:
          hours = ''

      ltd = ws.cell(row=wsRow,column=wsCol+10).value
      if (ltd == None):
        ltd = ''

      notes = ws.cell(row=wsRow,column=wsCol+11).value
      if (notes == None):
        notes = ''
 
      # TODO: if blankFlg == False, then make sure everything else is blank

      if (not blankFlg):
        self.entries.append(TSEntry(curDate,partA,partB,partC,partD,hours,ltd,notes))

      if (day.startswith('Sunday')):
        sundayFlg = True
      if (sundayFlg):
        if (blankFlg):
          sundayCnt += 1
        else:
          sundayCnt = 0;
      if (sundayCnt > 5):
        break

      wsRow += 1
      blankFlg = False

  #---------------------------------------------------------------------
  def Log(self):
    for i,entry in enumerate(self.entries):
      fullname = self.ssdata.fae.fullname.ljust(25)
      date  = str(self.entries[i].date)
      partA = str(self.entries[i].code).ljust(50)
      partB = str(self.entries[i].location).ljust(60)
      partC = str(self.entries[i].activity).ljust(45)
      partD = str(self.entries[i].product).ljust(35)
      hours = str(self.entries[i].hours).ljust(5)
      ltd   = str(self.entries[i].worktype).ljust(10)
      logging.debug(fullname + '|' + str(date) + '|' + partA + '|' + partB + '|' + partC + '|' + partD + '|' + hours + '|' + ltd)

